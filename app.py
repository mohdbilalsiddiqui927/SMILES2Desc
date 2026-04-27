import os
import subprocess
import tempfile
import platform
import io
from datetime import datetime

import streamlit as st
import pandas as pd
from rdkit import Chem, rdBase
from rdkit.Chem import Descriptors, Draw
from rdkit.Chem import SDWriter
from mordred import Calculator, descriptors as mordred_descriptors
import mordred
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment


# ══════════════════════════════════════════════════════
# PAGE CONFIG
# ══════════════════════════════════════════════════════
st.set_page_config(
    page_title="SMILES2Desc",
    page_icon="⬡",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ══════════════════════════════════════════════════════
# CUSTOM CSS
# ══════════════════════════════════════════════════════
st.markdown("""

<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Mono:wght@400;500&family=Syne:wght@700;800&family=DM+Sans:wght@300;400;500&display=swap');

:root {
    --bg:       #0d1117;
    --surface:  #161b22;
    --border:   #21262d;
    --accent:   #00d4a0;
    --accent2:  #0ea5e9;
    --warn:     #f59e0b;
    --danger:   #ef4444;
    --text:     #e6edf3;
    --muted:    #ffffff;
    --mono:     'DM Mono', monospace;
    --display:  'Syne', sans-serif;
    --body:     'DM Sans', sans-serif;
}

html, body, [class*="css"] {
    font-family: var(--body);
    background-color: var(--bg) !important;
    color: var(--text) !important;
}

#MainMenu, footer, header { visibility: hidden; }
.block-container { padding-top: 2rem !important; max-width: 1200px; }

.s2d-header {
    background: linear-gradient(135deg, #0d1117 0%, #161b22 60%, #0d1f1a 100%);
    border: 1px solid var(--border);
    border-left: 4px solid var(--accent);
    border-radius: 12px;
    padding: 2rem 2.5rem;
    margin-bottom: 2rem;
    position: relative;
    overflow: hidden;
}
.s2d-header::before {
    content: "⬡⬡⬡⬡⬡⬡⬡⬡⬡⬡⬡⬡⬡⬡⬡⬡⬡⬡⬡⬡";
    position: absolute;
    right: -10px; top: -10px;
    font-size: 2.2rem;
    opacity: 0.035;
    letter-spacing: 6px;
    line-height: 1.5;
    pointer-events: none;
    color: var(--accent);
}
.s2d-title {
    font-family: var(--display);
    font-size: 2.8rem;
    font-weight: 800;
    color: var(--text);
    letter-spacing: -1px;
    margin: 0 0 0.3rem 0;
    line-height: 1;
}
.s2d-title span { color: var(--accent); }
.s2d-subtitle {
    font-size: 0.92rem;
    color: var(--muted);
    font-weight: 300;
    margin: 0.4rem 0 0 0;
    letter-spacing: 0.3px;
}
.s2d-badge {
    display: inline-block;
    background: rgba(0,212,160,0.08);
    border: 1px solid rgba(0,212,160,0.25);
    color: var(--accent);
    font-family: var(--mono);
    font-size: 0.68rem;
    padding: 3px 12px;
    border-radius: 20px;
    margin-top: 1rem;
    letter-spacing: 1.5px;
    text-transform: uppercase;
}

.s2d-section {
    font-family: var(--mono);
    font-size: 0.65rem;
    letter-spacing: 2.5px;
    text-transform: uppercase;
    color: var(--accent);
    border-bottom: 1px solid var(--border);
    padding-bottom: 0.5rem;
    margin: 1.8rem 0 1rem 0;
}

.metric-grid {
    display: grid;
    grid-template-columns: repeat(4, 1fr);
    gap: 1rem;
    margin: 1.2rem 0 1.8rem 0;
}
.metric-card {
    background: var(--surface);
    border: 1px solid var(--border);
    border-radius: 10px;
    padding: 1.3rem 1.5rem;
    position: relative;
    overflow: hidden;
}
.metric-card::after {
    content: '';
    position: absolute;
    top: 0; left: 0; right: 0;
    height: 2px;
    background: var(--accent);
    border-radius: 10px 10px 0 0;
}
.metric-card.blue::after   { background: var(--accent2); }
.metric-card.warn::after   { background: var(--warn); }
.metric-card.danger::after { background: var(--danger); }
.metric-value {
    font-family: var(--display);
    font-size: 2.2rem;
    font-weight: 800;
    color: var(--text);
    line-height: 1;
    margin-bottom: 0.4rem;
}
.metric-label {
    font-family: var(--mono);
    font-size: 0.68rem;
    color: var(--muted);
    text-transform: uppercase;
    letter-spacing: 1.2px;
}

[data-testid="stFileUploader"] {
    background: #161b22 !important;
    border: 1.5px dashed #21262d !important;
    border-radius: 10px !important;
    color: #ffffff !important;
}
/* Fix inner uploader text + drag area */
[data-testid="stFileUploader"] * {
    color: #e6edf3 !important;
}

/* Remove light background inside */
[data-testid="stFileUploader"] section {
    background-color: #161b22 !important;
}
[data-testid="stFileUploader"]:hover {
    border-color: rgba(0,212,160,0.5) !important;
}

.stButton > button {
    background: var(--accent) !important;
    color: #0d1117 !important;
    font-family: var(--mono) !important;
    font-weight: 500 !important;
    font-size: 0.82rem !important;
    letter-spacing: 1.5px !important;
    text-transform: uppercase !important;
    border: none !important;
    border-radius: 8px !important;
    padding: 0.65rem 2rem !important;
    transition: all 0.2s ease !important;
}
.stButton > button:hover {
    background: #00f0b5 !important;
    transform: translateY(-1px) !important;
    box-shadow: 0 6px 24px rgba(0,212,160,0.25) !important;
}

.stDownloadButton > button {
    background: transparent !important;
    color: var(--accent) !important;
    border: 1px solid rgba(0,212,160,0.4) !important;
    font-family: var(--mono) !important;
    font-size: 0.78rem !important;
    letter-spacing: 1px !important;
    border-radius: 8px !important;
    transition: all 0.2s ease !important;
}
.stDownloadButton > button:hover {
    background: rgba(0,212,160,0.08) !important;
    border-color: var(--accent) !important;
    box-shadow: 0 0 16px rgba(0,212,160,0.15) !important;
}

[data-testid="stSidebar"] {
    background: var(--surface) !important;
    border-right: 1px solid var(--border) !important;
}
.sidebar-logo {
    font-family: var(--display);
    font-size: 1.5rem;
    font-weight: 800;
    color: var(--text);
    padding: 0.5rem 0 0.2rem 0;
    letter-spacing: -0.5px;
}
.sidebar-logo span { color: var(--accent); }
.sidebar-tagline {
    font-family: var(--mono);
    font-size: 0.62rem;
    color: var(--muted);
    letter-spacing: 1px;
    text-transform: uppercase;
    margin-bottom: 0.5rem;
}
.sidebar-section {
    font-family: var(--mono);
    font-size: 0.62rem;
    letter-spacing: 2.5px;
    text-transform: uppercase;
    color: var(--accent);
    margin: 1.2rem 0 0.6rem 0;
    padding-bottom: 0.35rem;
    border-bottom: 1px solid var(--border);
}

[data-testid="stExpander"] {
    background: var(--surface) !important;
    border: 1px solid var(--border) !important;
    border-radius: 10px !important;
    margin-bottom: 0.75rem !important;
}
[data-testid="stExpander"] summary {
    font-family: var(--mono) !important;
    font-size: 0.82rem !important;
    color: var(--muted) !important;
}

.stProgress > div > div { background: var(--accent) !important; }

[data-baseweb="tag"] {
    background: #0d1f1a !important;
    border: 1px solid #00d4a0 !important;
    color: #00d4a0 !important;
    font-family: var(--mono) !important;
    font-size: 0.73rem !important;
}
[data-baseweb="select"] > div {
    background-color: #161b22 !important;
}
.info-pill {
    display: inline-flex;
    align-items: center;
    gap: 8px;
    background: rgba(14,165,233,0.08);
    border: 1px solid rgba(14,165,233,0.2);
    color: #7dd3fc;
    font-family: var(--mono);
    font-size: 0.78rem;
    padding: 6px 14px;
    border-radius: 6px;
    margin: 0.5rem 0 1rem 0;
}

.pipeline-row {
    display: flex;
    gap: 0.6rem;
    margin: 0.5rem 0 1.5rem 0;
    flex-wrap: wrap;
}
.pipeline-chip {
    display: inline-flex;
    align-items: center;
    gap: 6px;
    background: var(--surface);
    border: 1px solid var(--border);
    border-radius: 6px;
    padding: 5px 14px;
    font-family: var(--mono);
    font-size: 0.75rem;
    color: var(--muted);
    letter-spacing: 0.5px;
}
.pipeline-chip.active {
    border-color: rgba(0,212,160,0.5);
    color: var(--accent);
    background: rgba(0,212,160,0.06);
}

.s2d-divider {
    border: none;
    border-top: 1px solid var(--border);
    margin: 2rem 0;
}

.sidebar-footer {
    font-family: var(--mono);
    font-size: 0.62rem;
    color: #4a5568;
    line-height: 1.8;
    margin-top: 1rem;
}
/* FORCE sidebar text to be visible */
[data-testid="stSidebar"] * {
    color: #ffffff !important;
    opacity: 1 !important;
}

/* specifically target checkbox + labels */
[data-testid="stSidebar"] .stCheckbox label span,
[data-testid="stSidebar"] .stCheckbox p,
[data-testid="stSidebar"] p {
    color: #ffffff !important;
    opacity: 1 !important;
}
/* Fix input fields text color (Max heavy atoms, JAR path, XML) */
[data-testid="stNumberInput"] input,
[data-testid="stTextInput"] input {
    color: #00d4a0 !important;        /* green text */
    background-color: #161b22 !important;  /* dark background */
}

/* Placeholder text (optional) */
[data-testid="stTextInput"] input::placeholder {
    color: #8b949e !important;
}
/* Browse files button styling */
[data-testid="stFileUploader"] button {
    background-color: #00d4a0 !important;   /* green */
    color: #0d1117 !important;              /* dark text */
    border-radius: 8px !important;
    border: none !important;
    font-weight: 600 !important;
}

/* Hover effect */
[data-testid="stFileUploader"] button:hover {
    background-color: #00f0b5 !important;
    color: #0d1117 !important;
}
/* Fix number input (+ / - buttons) */
[data-testid="stNumberInput"] button {
    background-color: #161b22 !important;  /* dark background */
    color: #00d4a0 !important;             /* green + - */
    border: 1px solid #00d4a0 !important;
}

/* Hover effect */
[data-testid="stNumberInput"] button:hover {
    background-color: #00d4a0 !important;
    color: #0d1117 !important;
}

/* Input box itself */
[data-testid="stNumberInput"] input {
    background-color: #161b22 !important;
    color: #00d4a0 !important;
}
</style>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════
# HEADER
# ══════════════════════════════════════════════════════
st.markdown("""
<div class="s2d-header">
    <div class="s2d-title">SMILES<span>2</span>Desc</div>
    <p class="s2d-subtitle">
        Molecular descriptor generation from SMILES strings &nbsp;·&nbsp;
        RDKit &nbsp;·&nbsp; Mordred &nbsp;·&nbsp; PaDEL
    </p>
    <span class="s2d-badge">v1.0 &nbsp;·&nbsp; </span>
</div>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("""
    <div class="sidebar-logo">S2<span>D</span></div>
    <div class="sidebar-tagline">SMILES2Desc · v1.0</div>
    """, unsafe_allow_html=True)
    st.markdown("---")

    st.markdown('<div class="sidebar-section">Descriptor Engines</div>', unsafe_allow_html=True)
    descriptor_types = st.multiselect(
        "engines",
        ["RDKit", "Mordred", "PaDEL"],
        default=["RDKit"],
        label_visibility="collapsed"
    )

    st.markdown('<div class="sidebar-section">Processing</div>', unsafe_allow_html=True)
    remove_duplicates    = st.checkbox("Remove duplicate SMILES",          value=True)
    remove_zero_variance = st.checkbox("Remove zero-variance descriptors", value=True)
    add_inchikey         = st.checkbox("Add InChIKey column",              value=True)
    show_preview         = st.checkbox("Show 2D structure preview",        value=True)
    max_heavy_atoms      = st.number_input(
        "Max heavy atoms (0 = unlimited)",
        min_value=0, max_value=500, value=100,
        help="Molecules exceeding this are skipped"
    )

    st.markdown('<div class="sidebar-section">Output Format</div>', unsafe_allow_html=True)
    output_format = st.radio(
        "fmt", ["CSV", "Excel (multi-sheet)"],
        label_visibility="collapsed"
    )

    if "PaDEL" in descriptor_types:
        st.markdown('<div class="sidebar-section">PaDEL / Java</div>', unsafe_allow_html=True)
        padel_jar = st.text_input(
            "JAR path",
            value=os.environ.get("PADEL_JAR_PATH", "./PaDEL-Descriptor.jar"),
            help="Full path to PaDEL-Descriptor.jar"
        )
        padel_xml = st.text_input(
            "Descriptor XML",
            value=os.environ.get("PADEL_XML_PATH", "./descriptors.xml")
        )
        padel_xmx = st.selectbox(
            "Java heap (Xmx)", ["1G", "2G", "4G"], index=1,
            help="Max RAM for Java process"
        )
    else:
        padel_jar = padel_xml = ""
        padel_xmx = "2G"

    st.markdown("---")
    st.markdown("""
    <div class="sidebar-footer">
        RDKit · Mordred · PaDEL<br>
        InChIKey · Canonical SMILES<br>
        Zero-variance filtering<br>
        Excel multi-sheet export
    </div>
    """, unsafe_allow_html=True)


# ══════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════
def check_java():
    r = subprocess.run(["java", "-version"], stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    return r.returncode == 0


def build_excel(final_df, metadata_dict, failed_list):
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Descriptors"
    hfill = PatternFill("solid", fgColor="0D2E26")
    hfont = Font(bold=True, color="00D4A0", name="Courier New")
    for row in dataframe_to_rows(final_df, index=False, header=True):
        ws.append(row)
    for cell in ws[1]:
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center")

    ws2 = wb.create_sheet("Metadata")
    ws2.append(["Key", "Value"])
    for cell in ws2[1]:
        cell.font = Font(bold=True, name="Courier New")
    for k, v in metadata_dict.items():
        ws2.append([k, str(v)])

    if failed_list:
        ws3 = wb.create_sheet("Failed SMILES")
        for row in dataframe_to_rows(pd.DataFrame(failed_list), index=False, header=True):
            ws3.append(row)
        for cell in ws3[1]:
            cell.font = Font(bold=True, color="EF4444")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ══════════════════════════════════════════════════════
# PIPELINE STATUS CHIPS
# ══════════════════════════════════════════════════════
chips_html = '<div class="pipeline-row">'
for eng in ["RDKit", "Mordred", "PaDEL"]:
    active = "active" if eng in descriptor_types else ""
    dot    = "⬡" if active else "○"
    chips_html += f'<span class="pipeline-chip {active}">{dot}&nbsp; {eng}</span>'
chips_html += '</div>'
st.markdown(chips_html, unsafe_allow_html=True)


# ══════════════════════════════════════════════════════
# FILE UPLOAD — read ONCE into session_state
# ══════════════════════════════════════════════════════
st.markdown('<div class="s2d-section">Input</div>', unsafe_allow_html=True)

uploaded_file = st.file_uploader(
    "Upload a CSV file with a **smiles** column. Additional columns are preserved in output.",
    type=["csv"],
    help="Max 200 MB · CSV format only"
)

if uploaded_file is not None:
    if st.session_state.get("_file_id") != uploaded_file.file_id:
        uploaded_file.seek(0)
        st.session_state["_df"]      = pd.read_csv(uploaded_file)
        st.session_state["_file_id"] = uploaded_file.file_id

    cached_df = st.session_state["_df"]
    st.markdown(
        f'<div class="info-pill">⬡ &nbsp;'
        f'<strong>{len(cached_df)}</strong> rows &nbsp;·&nbsp; '
        f'columns: <strong>{", ".join(cached_df.columns.tolist())}</strong>'
        f'</div>',
        unsafe_allow_html=True
    )


# ══════════════════════════════════════════════════════
# RUN BUTTON
# ══════════════════════════════════════════════════════
col_btn, _ = st.columns([1, 4])
with col_btn:
    run = st.button("⬡  Run Pipeline", type="primary", use_container_width=True)

if uploaded_file is not None and run:

    if not descriptor_types:
        st.error("Select at least one descriptor engine in the sidebar.")
        st.stop()

    df = st.session_state["_df"].copy()

    if "smiles" not in df.columns:
        st.error("CSV must contain a column named **smiles**.")
        st.stop()

    # ── Deduplicate ──
    if remove_duplicates:
        before = len(df)
        df     = df.drop_duplicates(subset="smiles").reset_index(drop=True)
        n_dupes = before - len(df)
        if n_dupes:
            st.warning(f"Removed **{n_dupes}** duplicate SMILES rows.")

    # ── Java preflight ──
    if "PaDEL" in descriptor_types:
        if not check_java():
            st.error("Java not found. PaDEL requires Java 8+.")
            st.stop()
        if not os.path.isfile(padel_jar):
            st.error(f"PaDEL JAR not found at: `{padel_jar}` — update path in sidebar.")
            st.stop()

    # ── SMILES validation ──
    st.markdown('<div class="s2d-section">Processing</div>', unsafe_allow_html=True)
    mols, valid_rows, failed = [], [], []
    prog = st.progress(0, text="Validating SMILES…")

    for idx, row in df.iterrows():
        smi = str(row["smiles"]).strip()
        mol = Chem.MolFromSmiles(smi)
        if mol is None:
            failed.append({"row": idx, "smiles": smi, "reason": "Invalid SMILES"})
            continue
        if max_heavy_atoms > 0 and mol.GetNumHeavyAtoms() > max_heavy_atoms:
            failed.append({"row": idx, "smiles": smi,
                           "reason": f"Exceeds {max_heavy_atoms} heavy atoms"})
            continue
        mols.append(mol)
        valid_rows.append(row)
        prog.progress((idx + 1) / len(df), text=f"Validating SMILES — {idx+1} / {len(df)}")

    prog.empty()

    if not mols:
        st.error("No valid SMILES found.")
        st.stop()

    base_df = pd.DataFrame(valid_rows).reset_index(drop=True)
    base_df["canonical_smiles"] = [Chem.MolToSmiles(m, canonical=True) for m in mols]

    if add_inchikey:
        from rdkit.Chem.inchi import MolToInchi
        from rdkit.Chem import rdinchi
        def safe_inchikey(mol):
            try:
                s = MolToInchi(mol)
                return rdinchi.InchiToInchiKey(s) or "N/A" if s else "N/A"
            except Exception:
                return "N/A"
        base_df["InChIKey"] = [safe_inchikey(m) for m in mols]

    result_dfs = [base_df]

    # ── RDKit ──
    if "RDKit" in descriptor_types:
        with st.spinner("RDKit — calculating descriptors…"):
            rows, p = [], st.progress(0, text="RDKit…")
            for i, mol in enumerate(mols):
                rows.append({f"RDKit_{n}": f(mol) for n, f in Descriptors.descList})
                p.progress((i + 1) / len(mols), text=f"RDKit — {i+1} / {len(mols)}")
            p.empty()
            rdk_df = pd.DataFrame(rows)
            result_dfs.append(rdk_df)
        st.success(f"RDKit complete — {rdk_df.shape[1]} descriptors")

    # ── Mordred ──
    if "Mordred" in descriptor_types:
        with st.spinner("Mordred — calculating descriptors…"):
            calc   = Calculator(mordred_descriptors, ignore_3D=True)
            mor_df = calc.pandas(mols)
            mor_df = mor_df.apply(pd.to_numeric, errors="coerce")
            mor_df = mor_df.add_prefix("Mordred_")
            result_dfs.append(mor_df)
        st.success(f"Mordred complete — {mor_df.shape[1]} descriptors")

    # ── PaDEL ──
    if "PaDEL" in descriptor_types:
        st.info("PaDEL is running — approximately 1 minute per 100 molecules.")
        with st.spinner("PaDEL — invoking Java CLI…"):
            try:
                with tempfile.TemporaryDirectory() as tmpdir:
                    for i, mol in enumerate(mols):
                        mol.SetProp("_Name", f"Mol_{i}")
                        w = SDWriter(os.path.join(tmpdir, f"mol_{i}.sdf"))
                        w.write(mol)
                        w.close()

                    out_csv = os.path.join(tmpdir, "padel_out.csv")
                    proc = subprocess.run(
                        ["java", "-Xms1G", f"-Xmx{padel_xmx}", "-jar", padel_jar,
                         "-descriptortypes", padel_xml, "-2d",
                         "-removesalt", "-standardizenitro",
                         "-fingerprints", "false",
                         "-dir", tmpdir, "-file", out_csv],
                        stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True
                    )
                    if proc.returncode != 0:
                        st.error("PaDEL exited with an error.")
                        with st.expander("PaDEL error details"):
                            st.code(proc.stderr)
                        st.stop()

                    pad_df = pd.read_csv(out_csv).add_prefix("PaDEL_")
                    pad_df = pad_df.drop(columns=["PaDEL_Name"], errors="ignore")
                    pad_df = pad_df.apply(pd.to_numeric, errors="coerce")
                    result_dfs.append(pad_df)
                st.success(f"PaDEL complete — {pad_df.shape[1]} descriptors")
            except Exception as e:
                st.error(f"PaDEL error: {e}")
                st.stop()

    # ── Merge ──
    final_df = pd.concat(result_dfs, axis=1)

    # ── Zero-variance removal ──
    if remove_zero_variance:
        desc_cols = [c for c in final_df.columns
                     if c.startswith(("RDKit_", "Mordred_", "PaDEL_"))]
        meta_cols = [c for c in final_df.columns if c not in desc_cols]
        keep      = [c for c in desc_cols if final_df[c].nunique() > 1]
        removed_zv = len(desc_cols) - len(keep)
        final_df   = pd.concat([final_df[meta_cols], final_df[keep]], axis=1)
        if removed_zv:
            st.info(f"Removed {removed_zv} zero-variance descriptor columns.")

    # ── Descriptor statistics ──
    desc_cols_final   = [c for c in final_df.columns
                         if c.startswith(("RDKit_", "Mordred_", "PaDEL_"))]
    numeric_desc_cols = (final_df[desc_cols_final]
                         .select_dtypes(include="number").columns.tolist())
    stats_df = None
    if numeric_desc_cols:
        stats_df = final_df[numeric_desc_cols].agg(["mean", "std", "min", "max"]).T
        stats_df["missing_%"] = (
            final_df[numeric_desc_cols].isna().mean() * 100
        ).round(2)
        stats_df = stats_df.round(4)

    # ── Metadata ──
    metadata = {
        "tool":                  "SMILES2Desc v1.0",
        "descriptor_types_used": ", ".join(descriptor_types),
        "rdkit_version":         rdBase.rdkitVersion,
        "mordred_version":       mordred.__version__ if "Mordred" in descriptor_types else "NA",
        "padel_version":         "PaDEL-Descriptor 2.21" if "PaDEL" in descriptor_types else "NA",
        "python_version":        platform.python_version(),
        "operating_system":      platform.platform(),
        "timestamp_utc":         datetime.utcnow().isoformat(),
        "input_molecules":       len(df),
        "successful_molecules":  len(final_df),
        "failed_molecules":      len(failed),
        "total_descriptors":     len(desc_cols_final),
        "zero_variance_removed": str(remove_zero_variance),
        "duplicates_removed":    str(remove_duplicates),
    }

    # ══════════════════════════════════════════════════════
    # RESULTS
    # ══════════════════════════════════════════════════════
    st.markdown('<hr class="s2d-divider">', unsafe_allow_html=True)
    st.markdown('<div class="s2d-section">Results</div>', unsafe_allow_html=True)

    st.markdown(f"""
    <div class="metric-grid">
        <div class="metric-card">
            <div class="metric-value">{len(final_df)}</div>
            <div class="metric-label">Molecules processed</div>
        </div>
        <div class="metric-card blue">
            <div class="metric-value">{len(desc_cols_final)}</div>
            <div class="metric-label">Descriptors generated</div>
        </div>
        <div class="metric-card warn">
            <div class="metric-value">{len(descriptor_types)}</div>
            <div class="metric-label">Engines used</div>
        </div>
        <div class="metric-card danger">
            <div class="metric-value">{len(failed)}</div>
            <div class="metric-label">Failed / skipped</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    if stats_df is not None:
        with st.expander("Descriptor Statistics  —  mean · std · min · max · missing %"):
            st.dataframe(stats_df, use_container_width=True)

    if show_preview and mols:
        with st.expander("2D Structure Preview  —  first 12 molecules"):
            img = Draw.MolsToGridImage(
                mols[:12],
                molsPerRow=4,
                subImgSize=(320, 260),
                legends=base_df["canonical_smiles"].tolist()[:12]
            )
            st.image(img, use_container_width=True)

    with st.expander("Table Preview  —  first 5 rows"):
        st.dataframe(final_df.head(), use_container_width=True)

    # ══════════════════════════════════════════════════════
    # DOWNLOADS
    # ══════════════════════════════════════════════════════
    st.markdown('<hr class="s2d-divider">', unsafe_allow_html=True)
    st.markdown('<div class="s2d-section">Download</div>', unsafe_allow_html=True)

    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")

    if output_format == "Excel (multi-sheet)":
        dl_col, note_col = st.columns([1, 2])
        with dl_col:
            st.download_button(
                "⬡  Download Excel",
                data=build_excel(final_df, metadata, failed),
                file_name=f"SMILES2Desc_{ts}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        with note_col:
            st.caption("3 sheets: **Descriptors** · **Metadata** · **Failed SMILES**")
    else:
        dl1, dl2, dl3 = st.columns(3)
        with dl1:
            st.download_button(
                "⬡  Descriptor CSV",
                data=final_df.to_csv(index=False),
                file_name=f"SMILES2Desc_descriptors_{ts}.csv",
                mime="text/csv",
                use_container_width=True
            )
        with dl2:
            st.download_button(
                "⬡  Metadata CSV",
                data=pd.DataFrame(list(metadata.items()),
                                   columns=["Key", "Value"]).to_csv(index=False),
                file_name=f"SMILES2Desc_metadata_{ts}.csv",
                mime="text/csv",
                use_container_width=True
            )

    if failed:
        st.markdown("")
        st.download_button(
            "⚠  Download Failed SMILES Report",
            data=pd.DataFrame(failed).to_csv(index=False),
            file_name=f"SMILES2Desc_failed_{ts}.csv",
            mime="text/csv"
        )
